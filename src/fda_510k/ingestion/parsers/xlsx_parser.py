from pathlib import Path

from openpyxl import load_workbook

from fda_510k.ingestion.parsers.base import BaseParser, ParsedPage, ParserResult


class XlsxParser(BaseParser):
    extensions = (".xlsx", ".xlsm")

    def parse(self, path: Path) -> ParserResult:
        wb = load_workbook(path, read_only=True, data_only=True)
        pages: list[ParsedPage] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None and str(c).strip()]
                if cells:
                    rows.append(" | ".join(cells))
            pages.append(ParsedPage(page_number=None, text=f"Sheet: {sheet_name}\n" + "\n".join(rows)))
        wb.close()
        return ParserResult(pages=pages)
